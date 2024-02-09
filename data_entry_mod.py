import csv
import openpyxl
from openpyxl import Workbook
from tkinter import *
from tkinter.ttk import Combobox
import pathlib
import re

class DataEntryForm:
    # Define class constants for error messages
    EMPTY_NAME_ERROR = "Name cannot be empty."
    INVALID_AGE_ERROR = "Invalid age. Please enter a positive integer."
    INVALID_CONTACT_ERROR = "Invalid contact number format. Please use xxx-xxx-xxxx."
    INVALID_GENDER_ERROR = "Invalid gender selection. Please choose Male or Female."
    ADDRESS_TOO_LONG_ERROR = "Address is too long. Please limit it to 100 characters."

    def __init__(self, root):
        self.root = root
        self.setup_gui()

    def setup_gui(self):
        # Add your GUI setup code here
        self.name1 = StringVar()
        self.age1 = StringVar()
        self.contact1 = StringVar()
       

        # ... (other GUI components)
        Label(root, text="Enter Details Here:", font=('Arial, bold', 20), bg="#C1CDC1", width=40, pady=5).pack()
        self.l1 = Label(self.root, text="Name:", font=('Arial,bold', 15), bg="#E0EEE0")
        self.l1.place(x=20, y=60)
        self.l2 = Label(self.root, text="Age:", font=('Arial,bold', 15), bg="#E0EEE0")
        self.l2.place(x=20, y=110)
        self.l3 = Label(self.root, text="Contact No.", font=('Arial,bold', 15), bg="#E0EEE0")
        self.l3.place(x=20, y=160)
        self.l4 = Label(self.root, text="Gender:", font=('Arial,bold', 15), bg="#E0EEE0")
        self.l4.place(x=320, y=110)
        self.l5 = Label(self.root, text="Address:", font=('Arial,bold', 15), bg="#E0EEE0")
        self.l5.place(x=20, y=210)
        self.error_label = Label(self.root, text="",fg='red', font=('Arial', 12))
        self.error_label.place(x=20, y=300)

        # self.error_label = Label(self.root, text="", fg="black")
        # self.error_label.pack()

        # self.bt1 = Button(self.root, text="Submit", command=self.submit)
        # self.bt1.pack()

        # self.bt2 = Button(self.root, text="Clear", command=self.clear)
        # self.bt2.pack()

        # self.bt3 = Button(self.root, text="Export Data", command=self.export_data)
        # self.bt3.pack()
        
        self.en1 = Entry(self.root, textvariable=self.name1, font=('Arial,bold', 15), width=30, bd=4)
        self.en1.place(x=130, y=60)
        self.en2 = Entry(self.root, textvariable=self.age1, font=('Arial,bold', 15), width=10, bd=4)
        self.en2.place(x=130, y=110)
        self.en3 = Entry(self.root, textvariable=self.contact1, font=('Arial,bold', 15), width=30, bd=4)
        self.en3.place(x=130, y=160)
        self.en5 = Text(self.root, width=41, height=5, bd=4)
        self.en5.place(x=130, y=210)

        self.c4 = Combobox(self.root, values=["Male", "Female"], font=('Arial', 12), width=8)
        self.c4.place(x=400, y=115)
        self.c4.set("Male")
        
        self.bt1 = Button(self.root, text="Submit", font=('Arial,bold', 15), bg="#C1CDC1", bd=5, command=self.submit)
        self.bt1.place(x=100, y=330)

        self.bt2 = Button(self.root, text="Clear", font=('Arial,bold', 15), bg="#C1CDC1", bd=5, command=self.clear)
        self.bt2.place(x=200, y=330)

        self.bt3 = Button(self.root, text="Exit", font=('Arial,bold', 15), bg="#C1CDC1", bd=5, command=lambda: root.destroy())
        self.bt3.place(x=300, y=330)

        self.bt4 = Button(self.root, text="Export", font=('Arial,bold', 15), bg="#C1CDC1", bd=5, command=lambda: self.export_data('csv'))
        self.bt4.place(x=400, y=330)
        
    def display_error(self, message):
        self.error_label.config(text=message, fg="red")
    
    def clear_errors(self):
        self.error_label.config(text="", fg="black")

    def submit(self):
        # Clear previous errors
        self.clear_errors()

        # Get user inputs
        name = self.name1.get()
        # Validation: Check for an empty name
        if not name.strip():
            self.display_error(self.EMPTY_NAME_ERROR)
            return  # Exit the function if there is an error

        try:
            age = int(self.age1.get())
            if age < 0:
                raise ValueError("Age must be a positive integer.")
        except ValueError:
            
            self.display_error(self.INVALID_AGE_ERROR)
            return  # Exit the function if there is an error

        contact = self.contact1.get()
        contact_pattern = re.compile(r'^\d{3}-\d{3}-\d{4}$')
        if not contact_pattern.match(contact):
            self.display_error(self.INVALID_CONTACT_ERROR)
            # return  # Exit the function if there is an error
    
        gender = self.c4.get()
        if gender not in ["Male", "Female"]:
            self.display_error("Invalid gender selection. Please choose Male or Female.")
            return  # Exit the function if there is an error
    
        address = self.en5.get(1.0, END)
        if len(address) > 100:
            # error_message = "Address is too long. Please limit it to 100 characters."
            self.display_error(self.ADDRESS_TOO_LONG_ERROR)
            return
        
        # Save data to Excel file
        self.save_to_excel(name, age, contact, gender, address)
    def save_to_excel(self, name, age, contact, gender, address):
        # File path
        file_path = pathlib.Path('DataEntry_file.xlsx')

        # Load or create workbook
        try:
            file = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            file = Workbook()
            sheet = file.active
            sheet["A1"] = "Name"
            sheet["B1"] = "Age"
            sheet["C1"] = "Contact No."
            sheet["D1"] = "Gender"
            sheet["E1"] = "Address"

        # Save data to Excel file
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=name)
        sheet.cell(column=2, row=sheet.max_row, value=age)
        sheet.cell(column=3, row=sheet.max_row, value=contact)
        sheet.cell(column=4, row=sheet.max_row, value=gender)
        sheet.cell(column=5, row=sheet.max_row, value=address)
        file.save(file_path)

    def export_data(self, format):
        # Get data to export
        name = self.name1.get()
        age = self.age1.get()
        contact = self.contact1.get()
        gender = self.c4.get()
        address = self.en5.get(1.0, END)
        data_to_export = [
            {"Name": name, "Age": age, "Contact": contact, "Gender": gender, "Address": address},
            # Add more data entries as needed
        ]

        # Export data to CSV
        self.export_to_csv(data_to_export, "exported_data.csv")

        # Export data to Excel
        self.export_to_excel(data_to_export, "exported_data.xlsx")
        print('exported')

        # Add other export formats as needed

    def export_to_csv(self, data, file_path):
        with open(file_path, mode='a', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=data[0].keys())
            
            # Write header only if the file is newly created
            if not file:
                writer.writeheader()

            writer.writerows(data)

        print(f"Data exported to {file_path}")

    def export_to_excel(self, data, file_path):
        workbook = Workbook()
        sheet = workbook.active

        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, entry in enumerate(data, 2):
            for col_num, value in enumerate(entry.values(), 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(file_path)
        print(f"Data exported to {file_path}")

    def clear(self):
        # Add your clear functionality here
        self.en1.delete(0, END)
        self.en2.delete(0, END)
        self.en3.delete(0, END)
        self.en5.delete(1.0, END)
        self.en1.focus()


# Example of usage
root = Tk()
root.geometry("550x400")
root.title("DataEntry In ExcelSheet")
root.config(bg="#E0EEE0")
data_entry_form = DataEntryForm(root)
root.mainloop()
